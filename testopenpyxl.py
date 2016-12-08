import openpyxl
import time

file = 'ABSi Bus Plan Rev N 090814.xlsx'

wb = openpyxl.load_workbook(filename=file,
                            read_only=True,
                            data_only=True
                            )


ws = wb["BS"]

t1 = time.time()

#for row in range(30):
#    for column in range(57):
#        read = ws.cell(row=row+1, column=column+1).value

t2 = time.time()

print("Time by cell is {}".format(str(t2-t1)))  # 9.84 seconds

t1 = time.time()

cells = ws['A1':'BE30']

for row in range(29):
    for column in range(56):
        value = cells[row+1][column+1].value


t2 = time.time()

print("Time by range is {}".format(str(t2-t1)))  # 0.024 seconds

# Note that time is spent on line 26 (cells = ws['A1':'BE30']) that actually
# loads the data.