"""
This class accesses data from excel data sheets.

Features to implement:
- Apply regular expressions to filter.
- Filter by cell range.
"""

from openpyxl import (load_workbook,
                      utils
                      )


class ExcelDB:

    def __init__(self, file):
        self.wb = load_workbook(filename=file,
                                read_only=True,
                                data_only=True
                                )
        self.wslist = self.wb.sheetnames

        filter = self.Filter() # What is this doing here?

    class Filter:
        """
        This class holds the filters applied to the excel sheet to determine
        the numbers that will be passed on for the Benford Law's test.
        There are three parameters: row labels, column labels, worksheets, and
        cell range. By default all these include any number cells in the
        worksheet but can be restricted to include or exclude certain cells
        based on those parameters.
        Note that a parameter is used if it is in the include list and not used
        if it is in the exclude list. If neither, the defaultinclude parameter
        will be used to determine whether or not to use the parameter. Ranges
        which are mistakenly both in the include and exclude lists will default
        to the defaultinclude setting.
        Row and column labels are strings, and are applied to any cell within
        the row or column.
        Cell ranges are numerical.
        All matches are exact, including string case.
        """

        class WorkSheets:

            include = set()

        class RowLabels:

            include = set()
            exclude = set()

            defaultinclude = False

        class ColLabels:

            include = set()
            exclude = set()

            defaultinclude = False

        class CellRange:

            include = set()
            exclude = set()

            defaultinclude = False

    def extractnumbers(self):
        """
        Apply the filters set in the Filter class and return the set of all
        numbers from the Excel file.

        :return: List of numbers from Excel file subject to Filter.
        """

        datareturn = []

        for worksheet in self.wslist:
            if worksheet in self.Filter.WorkSheets.include:

                ws = self.wb[worksheet]

                lastcell = (utils.get_column_letter(ws.max_column) +
                            str(ws.max_row)
                            )

                cells = ws['A1': lastcell]

                # Find range of cells that conform to Filter

                # Filter rows
                rows = []
                for idx in range(ws.max_row):
                    rw = cells[idx]
                    rwcells = set([cell.value for cell in rw])

                    ifincl = any(rwcells & self.Filter.RowLabels.include)
                    ifexcl = any(rwcells & self.Filter.RowLabels.exclude)
                    ifdefault = self.Filter.RowLabels.defaultinclude

                    if ifexcl is False and any((ifincl, ifdefault)):
                        rows += [idx]

                # Filter columns
                columns = []

                for idx in range(ws.max_column):
                    cl = tuple(row[idx] for row in cells)
                    clcells = set([cell.value for cell in cl])

                    ifincl = any(set(rwcells) & self.Filter.ColLabels.include)
                    ifexcl = any(set(rwcells) & self.Filter.ColLabels.exclude)
                    ifdefault = self.Filter.ColLabels.defaultinclude

                    if ifexcl is False and any((ifincl, ifdefault)):
                        columns += [idx]

                # Extract data from Excel sheet
                for rw in rows:
                    for cl in columns:
                        value = cells[rw][cl].value
                        if isinstance(value, (int, float)) and value != 0:
                            datareturn += [value]

        return datareturn
